from openpyxl import load_workbook, Workbook

class Stats:
    
    def __init__(self, inputFile='series.xlsx', outputName='output.txt'):
        # imported exel file
        inwb = load_workbook(inputFile +  ('.xlsx' if inputFile.find('.') < 0 else '') )
        # output exel file name
        self.outputName = outputName +  ('.txt' if outputName.find('.') < 0 else '')
        # creating output obj
        self.owb = open(self.outputName,"w+")
        # choosing sheet
        self.sheet = inwb.active

        self.lastI = 0
        self.lastJ = 0

    # calculating evrey cell in one row
    def calInRow(self, rowNum):

        for i in range(1, self.rowLen(rowNum)):
            alpha = 1
            # print('**' , self.rowLen(rowNum))
            # save somthing in the row!
            self.saveInCol(rowNum, i, i+1)
            self.saveInCol(rowNum, i+1, i)

            for j in range(2, self.rowLen(rowNum)-i): 
                criterio  = self.sheet.cell(rowNum, i+alpha).value
                pendiente = self.sheet.cell(rowNum, i+j).value
                if (criterio >= self.sheet.cell(rowNum, i).value):
                    break
                elif (criterio < pendiente):

                    self.saveInCol(rowNum, i, i+j)
                    self.saveInCol(rowNum, i+j, i)

                    alpha = j

        self.lastJ = i
        if criterio < pendiente:
            self.lastI = i+j
        else:
            self.lastI = i+1

        


    # saving in a cell
    def saveInCol(self, rowNum, i, j):
        self.owb.write("[{} {} {}]\n".format(i + self.lastI, j + self.lastJ, rowNum))

    # save the output
    def save(self):
        self.owb.close()
    
    def rowLen(self, rowNum):
        rowlen = len(self.sheet[rowNum])

        while (self.getCol(rowNum, rowlen) == None):
            rowlen = rowlen - 1 
        
        return rowlen

    def calInAllRows(self):
        for i in range(1, len(tuple((self.sheet.rows)))+1):
            self.calInRow(i)
    
    def getCol(self, rowNum, colNum):
        return self.sheet.cell(rowNum, colNum).value
